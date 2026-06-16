#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
뉴욕꼬맹이 전사지 자동 검색 & 열기 (엑셀 버전)
엑셀 파일에서 전사지 필요 목록을 읽고 파일을 자동으로 엽니다.

사용법:
1. Google Sheets의 데이터를 엑셀에 복사
2. transfer_list.xlsx 로 저장
3. 이 스크립트 실행
"""

import os
import sys
import subprocess
from pathlib import Path
from typing import List, Set

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl 라이브러리 필요. 설치:")
    print("   pip install openpyxl")
    sys.exit(1)


class TransferFileFinder:
    def __init__(self, excel_file: str, transfer_folders, file_ext: str = ".ai", batch_size: int = 3):
        """
        Args:
            excel_file: 엑셀 파일 경로 (transfer_list.xlsx)
            transfer_folders: 전사지 파일 폴더 (문자열 또는 리스트)
            file_ext: 파일 확장자 (.ai)
        """
        self.excel_file = Path(excel_file)

        # 단일 폴더면 리스트로 변환
        if isinstance(transfer_folders, str):
            self.transfer_folders = [Path(transfer_folders)]
        else:
            self.transfer_folders = [Path(f) for f in transfer_folders]

        self.file_ext = file_ext.lower()
        self.batch_size = batch_size

        if not self.excel_file.exists():
            raise FileNotFoundError(f"❌ 엑셀 파일을 찾을 수 없습니다: {excel_file}\n\n"
                                   f"📌 방법:\n"
                                   f"1. Google Sheets의 '전사지출력목록' 데이터 복사\n"
                                   f"2. 엑셀에 붙여넣기\n"
                                   f"3. {self.excel_file} 로 저장")

        for folder in self.transfer_folders:
            if not folder.exists():
                raise FileNotFoundError(f"❌ 폴더를 찾을 수 없습니다: {folder}")

    def get_print_needed_codes(self) -> Set[str]:
        """엑셀 파일에서 부족수량 > 0인 코드들 추출"""
        try:
            needed_codes = set()

            wb = openpyxl.load_workbook(self.excel_file, data_only=True)

            # 시트 찾기 (첫 번째 시트 사용)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            print(f"📋 '{sheet_name}' 시트 읽는 중...\n")

            # 헤더 읽기 (첫 번째 행)
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[col_idx] = str(cell.value).strip()

            print(f"헤더: {list(headers.values())}")

            # 열 위치 찾기
            code_col = None
            needs_col = None

            for col_idx, header in headers.items():
                if header == '전사지코드':
                    code_col = col_idx
                elif header == '부족수량':
                    needs_col = col_idx

            if not code_col or not needs_col:
                print(f"❌ 필수 열이 없습니다")
                print(f"필요: '전사지코드', '부족수량'")
                return set()

            # 데이터 읽기 (2행부터)
            for row_idx in range(2, ws.max_row + 1):
                code_cell = ws.cell(row_idx, code_col).value
                needs_cell = ws.cell(row_idx, needs_col).value

                if not code_cell:
                    continue

                code = str(code_cell).strip()

                try:
                    needs = int(needs_cell) if needs_cell else 0
                    if needs > 0:
                        needed_codes.add(code)
                        print(f"  🔴 {code}: {needs}개 필요")
                except (ValueError, TypeError):
                    pass

            print(f"\n📍 총 {len(needed_codes)}개 파일 검색\n")
            return needed_codes

        except Exception as e:
            print(f"❌ 엑셀 읽기 실패: {e}")
            raise

    def find_files(self, codes: Set[str]) -> dict:
        """여러 폴더에서 파일 검색 (하위폴더 포함)"""
        found = []
        not_found = []

        for code in sorted(codes):
            patterns = [
                f"{code}{self.file_ext}",
                f"{code.strip()}{self.file_ext}",
                f"{code.replace(' ', '')}{self.file_ext}",
            ]

            found_file = None

            # 모든 폴더에서 검색
            for folder in self.transfer_folders:
                for pattern in patterns:
                    # 하위폴더까지 검색 (** = 모든 하위폴더)
                    matching = list(folder.glob(f"**/{pattern}"))
                    if matching:
                        found_file = matching[0]
                        break
                if found_file:
                    break

            # 파일 대소문자 무시 검색 (하위폴더 포함)
            if not found_file:
                for folder in self.transfer_folders:
                    for file in folder.glob(f"**/*{self.file_ext}"):
                        if file.stem.upper() == code.upper():
                            found_file = file
                            break
                    if found_file:
                        break

            if found_file:
                found.append(str(found_file))
                print(f"  ✅ {code}: {found_file.name}")
            else:
                not_found.append(code)
                print(f"  ❌ {code}: 파일 없음")

        return {'found': found, 'not_found': not_found}

    def open_files(self, file_paths: List[str], batch_size: int = 3):
        """파일들을 배치로 나누어 열기 (PC 성능 고려)"""
        if not file_paths:
            print("열 파일이 없습니다")
            return

        print(f"\n📂 {len(file_paths)}개 파일을 {batch_size}개씩 열기...\n")

        for i, filepath in enumerate(file_paths):
            try:
                if sys.platform == 'win32':
                    os.startfile(filepath)
                elif sys.platform == 'darwin':
                    subprocess.Popen(['open', filepath])
                else:
                    subprocess.Popen(['xdg-open', filepath])
                print(f"  [{i+1}/{len(file_paths)}] ✅ {Path(filepath).name}")

                # 배치 단위마다 대기 (시스템 안정성)
                if (i + 1) % batch_size == 0:
                    print(f"\n⏸️  {batch_size}개 파일 열음. 다음을 계속하려면 엔터...")
                    input()
                    print()
            except Exception as e:
                print(f"  [{i+1}/{len(file_paths)}] ❌ {Path(filepath).name}: {e}")

    def run(self):
        """전체 프로세스 실행"""
        print("=" * 60)
        print("🔍 뉴욕꼬맹이 전사지 자동 검색")
        print("=" * 60 + "\n")

        # 1. 엑셀에서 인쇄 필요 항목 추출
        codes = self.get_print_needed_codes()
        if not codes:
            print("📌 인쇄 필요한 항목이 없습니다")
            return

        # 2. 파일 검색
        results = self.find_files(codes)

        # 3. 파일 열기
        if results['found']:
            self.open_files(results['found'], self.batch_size)

        # 결과 요약
        print("\n" + "=" * 60)
        print(f"✅ 완료: {len(results['found'])}개 파일 열음")
        if results['not_found']:
            print(f"⚠️  찾지 못한 파일: {', '.join(results['not_found'])}")
        print("=" * 60)


def main():
    # ⚙️ 설정
    EXCEL_FILE = "transfer_list.xlsx"  # Google Sheets 복사 → 엑셀에 붙여넣기 → 저장

    # 📁 전사지 파일 위치 (여러 폴더 지정 가능)
    TRANSFER_FOLDERS = [
        r"D:\전사지",
        # 회사에서 추가 폴더 확인 후 아래에 추가하세요:
        # r"폴더경로2",
        # r"폴더경로3",
    ]

    FILE_EXT = ".ai"   # 파일 확장자
    BATCH_SIZE = 3     # 한 번에 열 파일 개수 (3개 추천)

    try:
        finder = TransferFileFinder(EXCEL_FILE, TRANSFER_FOLDERS, FILE_EXT, BATCH_SIZE)
        finder.run()
    except Exception as e:
        print(f"\n❌ 오류: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
